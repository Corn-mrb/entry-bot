import os
from io import BytesIO
from datetime import datetime

from fastapi import FastAPI, Request, Query, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from database import (
    verify_token, get_stores, get_store, get_store_visits,
    get_all_visits_for_export, get_daily_stats, get_store_stats, _now_kst
)

# ----------------------------
# App Setup
# ----------------------------
app = FastAPI(title="Entry Bot Dashboard")
templates = Jinja2Templates(directory="templates")

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# ----------------------------
# 토큰 검증 헬퍼
# ----------------------------
def check_token(token: str):
    if not token:
        raise HTTPException(status_code=401, detail="토큰이 필요합니다.")

    token_data = verify_token(token)
    if not token_data:
        raise HTTPException(status_code=401, detail="유효하지 않거나 만료된 토큰입니다.")

    return token_data

# ----------------------------
# 대시보드 페이지
# ----------------------------
@app.get("/dashboard", response_class=HTMLResponse)
@limiter.limit("30/minute")
async def dashboard(request: Request, token: str = Query(None)):
    token_data = check_token(token)

    stores = get_stores()

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "token": token,
        "user": token_data,
        "stores": stores
    })

# ----------------------------
# API: 매장 목록
# ----------------------------
@app.get("/api/stores")
@limiter.limit("60/minute")
async def api_stores(request: Request, token: str = Query(None)):
    check_token(token)

    stores = get_stores()
    result = []

    for code, store in stores.items():
        visits = get_store_visits(code)
        result.append({
            "code": code,
            "name": store.get("store_name", ""),
            "visit_count": len(visits)
        })

    return {"stores": result}

# ----------------------------
# API: 방문 기록
# ----------------------------
@app.get("/api/visits")
@limiter.limit("60/minute")
async def api_visits(request: Request, token: str = Query(None), store_code: str = Query(None)):
    check_token(token)

    if store_code:
        store = get_store(store_code)
        if not store:
            raise HTTPException(status_code=404, detail="매장을 찾을 수 없습니다.")

        visits = get_store_visits(store_code)
        visits_data = []
        for v in visits:
            visits_data.append({
                "store_name": store.get("store_name", ""),
                "username": v.get("username", ""),
                "nickname": v.get("nickname", ""),
                "visit_date": v.get("visit_date", ""),
                "visit_time": v.get("visit_time", "")
            })
    else:
        visits_data = get_all_visits_for_export()

    return {"visits": visits_data}

# ----------------------------
# API: 일별 통계
# ----------------------------
@app.get("/api/stats/daily")
@limiter.limit("60/minute")
async def api_daily_stats(request: Request, token: str = Query(None), store_code: str = Query(None), days: int = Query(30)):
    check_token(token)

    stats = get_daily_stats(store_code, days)
    return {"stats": stats}

# ----------------------------
# API: 방문자별 통계
# ----------------------------
@app.get("/api/stats/visitors")
@limiter.limit("60/minute")
async def api_visitor_stats(request: Request, token: str = Query(None), store_code: str = Query(None)):
    check_token(token)

    if store_code:
        stats = get_store_stats(store_code)
    else:
        # 전체 매장 통계
        all_stats = {}
        for code in get_stores().keys():
            for stat in get_store_stats(code):
                uid = stat["user_id"]
                if uid not in all_stats:
                    all_stats[uid] = stat.copy()
                else:
                    all_stats[uid]["count"] += stat["count"]
        stats = sorted(all_stats.values(), key=lambda x: x["count"], reverse=True)

    return {"stats": stats}

# ----------------------------
# 내보내기: CSV
# ----------------------------
@app.get("/api/export/csv")
@limiter.limit("5/minute")
async def export_csv(request: Request, token: str = Query(None), store_code: str = Query(None)):
    check_token(token)

    if store_code:
        store = get_store(store_code)
        visits = get_store_visits(store_code)
        visits_data = []
        store_name = store.get("store_name", store_code) if store else store_code
        for v in visits:
            vd = v.get("visit_date", "")
            vt = v.get("visit_time", "")
            visits_data.append({
                "store_name": store_name,
                "nickname": v.get("nickname", ""),
                "user_id": v.get("user_id", ""),
                "visit_datetime": f"{vd} {vt}".strip(),
            })
        visits_data.sort(key=lambda x: x["visit_datetime"], reverse=True)
    else:
        visits_data = get_all_visits_for_export()

    # CSV 생성
    import csv
    output = BytesIO()

    # UTF-8 BOM for Excel
    output.write(b'\xef\xbb\xbf')

    import io
    text_output = io.StringIO()
    writer = csv.writer(text_output)
    writer.writerow(["장소", "닉네임", "유저아이디", "방문시각"])

    for v in visits_data:
        writer.writerow([
            v.get("store_name", ""),
            v.get("nickname", ""),
            v.get("user_id", ""),
            v.get("visit_datetime", ""),
        ])

    output.write(text_output.getvalue().encode('utf-8'))
    output.seek(0)

    filename = f"visits_{_now_kst().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ----------------------------
# 내보내기: Excel
# ----------------------------
@app.get("/api/export/xlsx")
@limiter.limit("5/minute")
async def export_xlsx(request: Request, token: str = Query(None), store_code: str = Query(None)):
    check_token(token)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 설치되지 않았습니다.")

    if store_code:
        store = get_store(store_code)
        visits = get_store_visits(store_code)
        visits_data = []
        store_name = store.get("store_name", store_code) if store else store_code
        for v in visits:
            vd = v.get("visit_date", "")
            vt = v.get("visit_time", "")
            visits_data.append({
                "store_name": store_name,
                "nickname": v.get("nickname", ""),
                "user_id": v.get("user_id", ""),
                "visit_datetime": f"{vd} {vt}".strip(),
            })
        visits_data.sort(key=lambda x: x["visit_datetime"], reverse=True)
    else:
        visits_data = get_all_visits_for_export()

    wb = Workbook()
    ws = wb.active
    ws.title = "방문 기록"

    # 헤더
    headers = ["장소", "닉네임", "유저아이디", "방문시각"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    for row, v in enumerate(visits_data, 2):
        ws.cell(row=row, column=1, value=v.get("store_name", ""))
        ws.cell(row=row, column=2, value=v.get("nickname", ""))
        ws.cell(row=row, column=3, value=v.get("user_id", ""))
        ws.cell(row=row, column=4, value=v.get("visit_datetime", ""))

    # 열 너비
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"visits_{_now_kst().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ----------------------------
# 내보내기: PDF
# ----------------------------
@app.get("/api/export/pdf")
@limiter.limit("3/minute")
async def export_pdf(request: Request, token: str = Query(None), store_code: str = Query(None)):
    check_token(token)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab 패키지가 설치되지 않았습니다.")

    if store_code:
        store = get_store(store_code)
        visits = get_store_visits(store_code)
        visits_data = []
        store_name = store.get("store_name", store_code) if store else store_code
        for v in visits:
            vd = v.get("visit_date", "")
            vt = v.get("visit_time", "")
            visits_data.append({
                "store_name": store_name,
                "nickname": v.get("nickname", ""),
                "user_id": v.get("user_id", ""),
                "visit_datetime": f"{vd} {vt}".strip(),
            })
        visits_data.sort(key=lambda x: x["visit_datetime"], reverse=True)
    else:
        visits_data = get_all_visits_for_export()

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)

    # 테이블 데이터
    table_data = [["장소", "닉네임", "유저아이디", "방문시각"]]
    for v in visits_data[:100]:  # 최대 100건
        table_data.append([
            v.get("store_name", "")[:15],
            v.get("nickname", "")[:15],
            str(v.get("user_id", "")),
            v.get("visit_datetime", ""),
        ])

    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    doc.build([table])
    output.seek(0)

    filename = f"visits_{_now_kst().strftime('%Y%m%d_%H%M%S')}.pdf"

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ----------------------------
# Health Check
# ----------------------------
@app.get("/health")
async def health():
    return {"status": "ok"}

# ----------------------------
# 실행
# ----------------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
