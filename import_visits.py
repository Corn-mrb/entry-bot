"""
다른봇 방문기록 → entry-bot visits.json 통합 스크립트
대상: visits_20260326_191221_KST.xlsx 전체기록 시트
매장코드: 94 (비트코인하우스오리진)
"""
import json
import openpyxl
from datetime import datetime

XLSX_PATH = '/Users/gim-yeongseog/Downloads/bots/visits_20260326_191221_KST.xlsx'
VISITS_PATH = '/Users/gim-yeongseog/Downloads/bots/entry-bot/data/visits.json'
STORE_CODE = '94'

# 기존 visits.json 로드
with open(VISITS_PATH, 'r', encoding='utf-8') as f:
    visits = json.load(f)

if STORE_CODE not in visits:
    visits[STORE_CODE] = []

# 중복 체크용 (user_id + visit_date)
existing = {(v['user_id'], v['visit_date']) for v in visits[STORE_CODE]}

# xlsx 읽기
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb['전체기록']

added = 0
skipped = 0

for r in range(2, ws.max_row + 1):
    place    = ws.cell(r, 1).value
    nickname = ws.cell(r, 2).value
    uid_raw  = ws.cell(r, 3).value
    dt_raw   = ws.cell(r, 4).value

    if not uid_raw or not dt_raw:
        continue

    user_id = int(uid_raw)
    dt_str = str(dt_raw)  # "YYYY-MM-DD HH:MM:SS"
    visit_date = dt_str[:10]
    visit_time = dt_str[11:19]
    created_at = f"{visit_date}T{visit_time}+09:00"

    key = (user_id, visit_date)
    if key in existing:
        skipped += 1
        continue

    visits[STORE_CODE].append({
        'user_id': user_id,
        'username': '',
        'nickname': nickname or '',
        'visit_date': visit_date,
        'visit_time': visit_time,
        'created_at': created_at
    })
    existing.add(key)
    added += 1

# 날짜/시간 순 정렬
visits[STORE_CODE].sort(key=lambda x: (x['visit_date'], x['visit_time']))

# 저장
with open(VISITS_PATH, 'w', encoding='utf-8') as f:
    json.dump(visits, f, ensure_ascii=False, indent=2)

print(f'추가: {added}건 / 중복 스킵: {skipped}건')
print(f'store 94 총 방문기록: {len(visits[STORE_CODE])}건')
